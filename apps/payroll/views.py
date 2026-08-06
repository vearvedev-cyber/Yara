"""
Payroll API ViewSets
"""

from decimal import Decimal
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, timedelta
import csv
from io import StringIO
from django.db.models import Q
from .models import (
    PayrollComponent, PayrollEntry, SalaryRange, TitleBreakdown, PayrollPeriod,
    Payslip, PayslipDeduction, PayslipAuditLog, PAYEReturn,
    WorkspaceStatutorySettings, PayeTaxBand, ComplianceDocument,
)
from .utils import get_tax_band_breakdown, calculate_gross_from_net
from .pdf_generator import generate_payslip_pdf
from .serializers import (
    PayrollComponentSerializer,
    PayrollEntrySerializer,
    SalaryRangeSerializer,
    TitleBreakdownSerializer,
    PayrollPeriodSerializer,
    PayslipSerializer,
    PayslipCreateSerializer,
    PayslipAuditLogSerializer,
    PAYEReturnSerializer,
    PAYEReturnDetailRowSerializer,
    WorkspaceStatutorySettingsSerializer,
    PayeTaxBandSerializer,
    ComplianceDocumentSerializer,
)


class PayrollComponentViewSet(viewsets.ModelViewSet):
    """ViewSet for position-based payroll components."""
    queryset = PayrollComponent.objects.all()
    serializer_class = PayrollComponentSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['position']
    ordering_fields = ['position']
    ordering = ['position']

    def perform_create(self, serializer):
        serializer.save()


class PayrollEntryViewSet(viewsets.ModelViewSet):
    """ViewSet for employee payroll entries."""
    queryset = PayrollEntry.objects.select_related('employee').all()
    serializer_class = PayrollEntrySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'currency']
    search_fields = ['employee__first_name', 'employee__last_name', 'employee__employee_number', 'department']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = super().get_queryset()
        # Filter by workspace context if available
        if hasattr(self.request, 'workspace') and self.request.workspace:
            qs = qs.filter(employee__workspace=self.request.workspace)
        return qs

    def perform_create(self, serializer):
        # Auto-assign workspace from request context
        if hasattr(self.request, 'workspace') and self.request.workspace:
            # Ensure employee belongs to this workspace
            employee = serializer.validated_data.get('employee')
            if employee and not employee.workspace_id:
                employee.workspace = self.request.workspace
                employee.save(update_fields=['workspace'])
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Creating payroll entry with validated_data: {serializer.validated_data}")
        
        serializer.save()


class SalaryRangeViewSet(viewsets.ModelViewSet):
    """Buckets of employees by salary range (EMPLOYEE SALARY RANGE section)."""

    queryset = SalaryRange.objects.all()
    serializer_class = SalaryRangeSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['label']
    ordering_fields = ['label', 'employee_count']
    ordering = ['label']

    def list(self, request, *args, **kwargs):
        """Calculate salary ranges dynamically from actual contract data"""
        from apps.hcm.models import Contract
        from .models import PayrollEntry
        
        # Get all active contracts with salary data
        active_contracts = Contract.objects.filter(status='ACTIVE').exclude(basic_salary__isnull=True)
        if hasattr(request, 'workspace') and request.workspace:
            active_contracts = active_contracts.filter(employee__workspace=request.workspace)

        use_payroll_entries = not active_contracts.exists()
        if use_payroll_entries:
            salary_qs = PayrollEntry.objects.all()
            salary_field = 'gross'
            if hasattr(request, 'workspace') and request.workspace:
                salary_qs = salary_qs.filter(employee__workspace=request.workspace)
        else:
            salary_qs = active_contracts
            salary_field = 'basic_salary'
        
        # Define salary ranges
        ranges = [
            {'label': '0K - 5K', 'min': 0, 'max': 5000},
            {'label': '5K - 10K', 'min': 5000, 'max': 10000},
            {'label': '10K - 15K', 'min': 10000, 'max': 15000},
            {'label': '15K - 20K', 'min': 15000, 'max': 20000},
            {'label': '20K - 30K', 'min': 20000, 'max': 30000},
            {'label': '30K - 40K', 'min': 30000, 'max': 40000},
            {'label': '40K - 50K', 'min': 40000, 'max': 50000},
            {'label': '50K+', 'min': 50000, 'max': 999999999},
        ]
        
        # Calculate employee counts for each range
        result = []
        for r in ranges:
            count = salary_qs.filter(
                **{
                    f'{salary_field}__gte': r['min'],
                    f'{salary_field}__lt': r['max']
                }
            ).count()
            result.append({
                'id': len(result) + 1,
                'label': r['label'],
                'min_gross': r['min'],
                'max_gross': r['max'],
                'employee_count': count,
                'currency': 'ZMW'
            })
        
        return Response(result)


class TitleBreakdownViewSet(viewsets.ModelViewSet):
    """Per-title pay breakdown (TITLE BREAKDOWN section)."""

    serializer_class = TitleBreakdownSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['position']
    ordering_fields = ['position', 'gross', 'net']
    ordering = ['position']
    
    def get_queryset(self):
        qs = TitleBreakdown.objects.all()
        if hasattr(self.request, 'workspace') and self.request.workspace:
            qs = qs.filter(employee__workspace=self.request.workspace)
        return qs


class PayrollPeriodViewSet(viewsets.ModelViewSet):
    """ViewSet for payroll periods (months)"""
    queryset = PayrollPeriod.objects.all()
    serializer_class = PayrollPeriodSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['year', 'month', 'status']
    ordering = ['-year', '-month']

    def perform_create(self, serializer):
        serializer.save()


class PayslipViewSet(viewsets.ModelViewSet):
    """ViewSet for employee payslips with Zambian statutory calculations"""
    queryset = Payslip.objects.select_related('employee', 'period').prefetch_related('custom_deductions').filter(is_active=True)
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'period', 'period__year', 'period__month', 'is_processed', 'is_active']
    search_fields = ['employee__first_name', 'employee__last_name', 'employee__employee_id']
    ordering = ['-period__year', '-period__month', 'employee__employee_id']
    
    def get_queryset(self):
        qs = super().get_queryset()
        # Filter by workspace through employee
        if hasattr(self.request, 'workspace') and self.request.workspace:
            qs = qs.filter(employee__workspace=self.request.workspace)
        return qs
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return PayslipCreateSerializer
        return PayslipSerializer
    
    def perform_create(self, serializer):
        payslip = serializer.save()
        # Create audit log
        PayslipAuditLog.objects.create(
            payslip=payslip,
            action='CREATED',
            user=self.request.user,
            changes={'initial_data': serializer.validated_data}
        )
    
    def perform_update(self, serializer):
        # Capture old values
        old_data = {
            'basic_salary': str(serializer.instance.basic_salary),
            'gross_salary': str(serializer.instance.gross_salary),
            'net_salary': str(serializer.instance.net_salary),
        }
        payslip = serializer.save()
        # Capture new values
        new_data = {
            'basic_salary': str(payslip.basic_salary),
            'gross_salary': str(payslip.gross_salary),
            'net_salary': str(payslip.net_salary),
        }
        # Create audit log
        PayslipAuditLog.objects.create(
            payslip=payslip,
            action='UPDATED',
            user=self.request.user,
            changes={'old': old_data, 'new': new_data}
        )
    
    def perform_destroy(self, instance):
        # Create audit log before deletion
        PayslipAuditLog.objects.create(
            payslip=instance,
            action='DELETED',
            user=self.request.user,
            changes={'deleted_data': {
                'employee': instance.employee.full_name,
                'period': str(instance.period),
                'net_salary': str(instance.net_salary)
            }}
        )
        instance.delete()

    def _ensure_calculated(self, payslip, user=None):
        if not payslip.is_processed:
            payslip.calculate()
            payslip.save()
            PayslipAuditLog.objects.create(
                payslip=payslip,
                action='CALCULATED',
                user=user,
                changes={'auto_calculated': True}
            )

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        items = page if page is not None else queryset
        for payslip in items:
            self._ensure_calculated(payslip, request.user)
        serializer = self.get_serializer(items, many=True)
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        payslip = self.get_object()
        self._ensure_calculated(payslip, request.user)
        serializer = self.get_serializer(payslip)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def recalculate(self, request, pk=None):
        """Recalculate payslip by creating a new version"""
        old_payslip = self.get_object()
        old_net = str(old_payslip.net_salary)
        
        # Deactivate old version
        old_payslip.is_active = False
        old_payslip.save()
        
        # Create new version
        new_payslip = Payslip.objects.create(
            employee=old_payslip.employee,
            period=old_payslip.period,
            basic_salary=old_payslip.basic_salary,
            housing_allowance=old_payslip.housing_allowance,
            transportation_allowance=old_payslip.transportation_allowance,
            lunch_allowance=old_payslip.lunch_allowance,
            other_allowances=old_payslip.other_allowances,
            overtime_payment=old_payslip.overtime_payment,
            bonus=old_payslip.bonus,
            unpaid_leave_days=old_payslip.unpaid_leave_days,
            unpaid_leave_deduction=old_payslip.unpaid_leave_deduction,
            absenteeism_days=old_payslip.absenteeism_days,
            absenteeism_deduction=old_payslip.absenteeism_deduction,
            gross_salary=old_payslip.gross_salary,
            napsa_employee=old_payslip.napsa_employee,
            napsa_employer=old_payslip.napsa_employer,
            paye_tax=old_payslip.paye_tax,
            nhima_employee=old_payslip.nhima_employee,
            nhima_employer=old_payslip.nhima_employer,
            total_custom_deductions=old_payslip.total_custom_deductions,
            total_deductions=old_payslip.total_deductions,
            net_salary=old_payslip.net_salary,
            version=old_payslip.version + 1,
            is_active=True,
            original_payslip=old_payslip.original_payslip if old_payslip.original_payslip else old_payslip,
            is_processed=old_payslip.is_processed,
            employer_borne_deductions=old_payslip.employer_borne_deductions,
            notes=old_payslip.notes,
        )

        # Copy custom deductions
        for deduction in old_payslip.custom_deductions.all():
            PayslipDeduction.objects.create(
                payslip=new_payslip,
                description=deduction.description,
                amount=deduction.amount
            )

        # Recalculate
        new_payslip.calculate()
        new_payslip.save()
        
        # Create audit log
        PayslipAuditLog.objects.create(
            payslip=new_payslip,
            action='CALCULATED',
            user=request.user,
            changes={
                'old_version': old_payslip.version,
                'new_version': new_payslip.version,
                'old_net': old_net,
                'new_net': str(new_payslip.net_salary)
            }
        )
        
        serializer = PayslipSerializer(new_payslip)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def tax_breakdown(self, request, pk=None):
        """Return PAYE band breakdown for this payslip using workspace-specific bands"""
        payslip = self.get_object()

        # Resolve workspace from employee memberships if available
        workspace = None
        memberships = getattr(payslip.employee, 'workspace_memberships', None)
        if memberships is not None and memberships.exists():
            workspace = memberships.first().workspace

        adjusted_gross = float(payslip.gross_salary) - float(payslip.unpaid_leave_deduction)
        chargeable_income = adjusted_gross - float(payslip.napsa_employee)

        breakdown = get_tax_band_breakdown(chargeable_income, workspace)

        return Response({
            'adjusted_gross': round(adjusted_gross, 2),
            'chargeable_income': round(chargeable_income, 2),
            'bands': breakdown,
        })
    
    @action(detail=True, methods=['get'])
    def audit_history(self, request, pk=None):
        """Get audit trail for this payslip"""
        payslip = self.get_object()
        logs = payslip.audit_logs.all()
        serializer = PayslipAuditLogSerializer(logs, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def version_history(self, request, pk=None):
        """Get all versions of this payslip"""
        payslip = self.get_object()
        
        # Get the original payslip (root of version chain)
        original = payslip.original_payslip if payslip.original_payslip else payslip
        
        # Get all versions including the original
        if original == payslip:
            # This is the original, get all its versions plus itself
            all_versions = [original] + list(original.versions.all())
        else:
            # This is a version, get original and all its versions
            all_versions = [original] + list(original.versions.all())
        
        # Sort by version number
        all_versions.sort(key=lambda p: p.version)
        
        serializer = PayslipSerializer(all_versions, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Bulk create payslips for all employees for a specific period
        POST: {
            "year": 2026,
            "month": 1,
            "employee_ids": [1, 2, 3]  // optional, if not provided creates for all payroll-enrolled employees
        }
        """
        from apps.hcm.models import Employee
        
        year = request.data.get('year')
        month = request.data.get('month')
        employee_ids = request.data.get('employee_ids')
        # Optional bulk override; None means "inherit from each employee's PayrollEntry"
        bulk_employer_borne = request.data.get('employer_borne_deductions', None)
        
        if not year or not month:
            return Response(
                {'error': 'year and month are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get or create period
        period, created = PayrollPeriod.objects.get_or_create(
            year=year,
            month=month
        )
        
        # Only generate payslips for employees that are in payroll (have PayrollEntry rows).
        payroll_entries_qs = PayrollEntry.objects.all()

        # Apply workspace filtering if available.
        if hasattr(request, 'workspace') and request.workspace:
            payroll_entries_qs = payroll_entries_qs.filter(employee__workspace=request.workspace)

        if employee_ids:
            payroll_entries_qs = payroll_entries_qs.filter(employee_id__in=employee_ids)

        payroll_employee_ids = payroll_entries_qs.values_list('employee_id', flat=True).distinct()
        employees = Employee.objects.filter(id__in=payroll_employee_ids)
        
        created_payslips = []
        errors = []
        skipped_existing = 0
        skipped_no_salary = 0
        skipped_other = 0
        
        # Import leave models for integration
        from apps.leave.models import LeaveRequest, SickNote, Absenteeism, DoubleTicketRequest
        from datetime import datetime
        
        # Get the date range for the payroll period
        period_start = datetime(year=int(year), month=int(month), day=1).date()
        if int(month) == 12:
            period_end = datetime(year=int(year)+1, month=1, day=1).date()
        else:
            period_end = datetime(year=int(year), month=int(month)+1, day=1).date()
        
        for employee in employees:
            # Check if payslip already exists
            if Payslip.objects.filter(employee=employee, period=period).exists():
                errors.append(f"Payslip already exists for {employee.full_name}")
                skipped_existing += 1
                continue
            
            # Try to get salary structure from PayrollEntry (Salary Management Dashboard)
            payroll_entry = employee.payroll_entries.order_by('-updated_at', '-created_at').first()
            
            if not payroll_entry:
                errors.append(f"No salary data found for {employee.full_name}. Please add employee to payroll first.")
                skipped_no_salary += 1
                continue
            
            # Use comprehensive salary structure from PayrollEntry
            basic_salary = payroll_entry.basic
            housing_allowance = payroll_entry.housing
            transportation_allowance = payroll_entry.transportation
            lunch_allowance = payroll_entry.lunch
            
            # Calculate unpaid leave days for this period
            unpaid_leave_days = 0
            unpaid_leaves = LeaveRequest.objects.filter(
                employee=employee,
                leave_type='UNPAID',
                status='APPROVED',
                start_date__lt=period_end,
                end_date__gte=period_start
            )
            for leave in unpaid_leaves:
                # Calculate overlap days
                overlap_start = max(leave.start_date, period_start)
                overlap_end = min(leave.end_date, period_end)
                if overlap_start <= overlap_end:
                    unpaid_leave_days += (overlap_end - overlap_start).days + 1
            
            # Calculate approved sick note days for this period (no deduction)
            sick_dates = set()

            sick_notes = SickNote.objects.filter(
                employee=employee,
                status='APPROVED',
                start_date__lt=period_end,
                end_date__gte=period_start
            )
            for note in sick_notes:
                overlap_start = max(note.start_date, period_start)
                overlap_end = min(note.end_date, period_end)
                if overlap_start <= overlap_end:
                    days = (overlap_end - overlap_start).days + 1
                    for i in range(days):
                        sick_dates.add(overlap_start + timedelta(days=i))

            sick_leaves = LeaveRequest.objects.filter(
                employee=employee,
                leave_type='SICK',
                status='APPROVED',
                start_date__lt=period_end,
                end_date__gte=period_start
            )
            for leave in sick_leaves:
                overlap_start = max(leave.start_date, period_start)
                overlap_end = min(leave.end_date, period_end)
                if overlap_start <= overlap_end:
                    days = (overlap_end - overlap_start).days + 1
                    for i in range(days):
                        sick_dates.add(overlap_start + timedelta(days=i))

            sick_note_days = len(sick_dates)

            # Calculate absenteeism days for this period (exclude approved sick notes)
            absences = Absenteeism.objects.filter(
                employee=employee,
                status='UNJUSTIFIED',
                date__gte=period_start,
                date__lt=period_end
            )
            absenteeism_days = sum(1 for absence in absences if absence.date not in sick_dates)
            
            # Calculate double ticket (Sunday/holiday work) payments
            double_ticket_amount = 0
            double_tickets = DoubleTicketRequest.objects.filter(
                employee=employee,
                status='APPROVED',
                work_date__gte=period_start,
                work_date__lt=period_end
            )
            for ticket in double_tickets:
                if ticket.calculated_amount:
                    double_ticket_amount += float(ticket.calculated_amount)
            
            # Create payslip with comprehensive data
            notes = []
            if unpaid_leave_days:
                notes.append(f"Unpaid leave days: {unpaid_leave_days}")
            if absenteeism_days:
                notes.append(f"Absenteeism days: {absenteeism_days}")
            if sick_note_days:
                notes.append(f"Sick days: {sick_note_days}")

            # Determine employer_borne: bulk override takes precedence; fallback to per-employee setting
            if bulk_employer_borne is not None:
                employer_borne = bool(bulk_employer_borne)
            else:
                employer_borne = bool(getattr(payroll_entry, 'employer_borne_deductions', False))

            payslip = Payslip.objects.create(
                employee=employee,
                period=period,
                basic_salary=basic_salary,
                housing_allowance=housing_allowance,
                transportation_allowance=transportation_allowance,
                lunch_allowance=lunch_allowance,
                other_allowances=0,
                overtime_payment=double_ticket_amount,  # Double ticket is a form of overtime
                unpaid_leave_days=unpaid_leave_days,
                absenteeism_days=absenteeism_days,
                gross_salary=0,
                total_deductions=0,
                net_salary=0,
                employer_borne_deductions=employer_borne,
                notes='; '.join(notes)
            )
            
            # Calculate all deductions and net
            payslip.calculate()
            payslip.save()
            
            created_payslips.append(payslip)
        
        serializer = PayslipSerializer(created_payslips, many=True)
        skipped_other = max(len(errors) - skipped_existing - skipped_no_salary, 0)
        return Response({
            'created': len(created_payslips),
            'errors': errors,
            'summary': {
                'target_employees': employees.count(),
                'skipped_existing': skipped_existing,
                'skipped_no_salary': skipped_no_salary,
                'skipped_other': skipped_other,
            },
            'payslips': serializer.data
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['post'])
    def bulk_set_employer_borne(self, request):
        """
        Switch existing payslips for a period between employer-borne and standard mode.

        Switching TO employer-borne (employer_borne_deductions=true):
          - The employee's current net_salary becomes the new gross_salary
            (they keep the same take-home; employer now covers deductions on top)
          - Salary components (basic/housing/transport/lunch) are recalculated
            from the new gross using workspace ratios
          - PayrollEntry is also updated so next month generates correctly

        Switching BACK to standard (employer_borne_deductions=false):
          - The current gross_salary is back-calculated to find the gross that
            produces net = current gross (i.e. reverses the operation)
        """
        from .utils import get_statutory_settings, calculate_gross_from_net
        from apps.hcm.models import Employee

        year = request.data.get('year')
        month = request.data.get('month')
        employer_borne = bool(request.data.get('employer_borne_deductions', False))

        if not year or not month:
            return Response({'error': 'year and month are required'}, status=status.HTTP_400_BAD_REQUEST)

        payslips = Payslip.objects.filter(
            period__year=year,
            period__month=month,
            is_active=True,
        ).select_related('employee', 'employee__workspace')
        if hasattr(request, 'workspace') and request.workspace:
            payslips = payslips.filter(employee__workspace=request.workspace)

        updated = 0
        for payslip in payslips:
            workspace = payslip.employee.workspace
            statutory = get_statutory_settings(workspace)
            basic_r = statutory['basic_ratio']
            housing_r = statutory['housing_ratio']
            transport_r = statutory['transport_ratio']
            lunch_r = statutory['lunch_ratio']

            if employer_borne:
                # The old net becomes the new gross — employee takes home exactly what they earned before
                new_gross = float(payslip.net_salary)
            else:
                # Reverting: the current gross_salary WAS the old net; find the gross that gives that net
                result = calculate_gross_from_net(float(payslip.gross_salary), workspace=workspace)
                new_gross = result['gross_salary']

            # Recalculate components from new gross using workspace ratios
            housing = round(new_gross * housing_r, 2)
            transport = round(new_gross * transport_r, 2)
            lunch = round(new_gross * lunch_r, 2)
            basic = round(new_gross - housing - transport - lunch, 2)

            payslip.basic_salary = basic
            payslip.housing_allowance = housing
            payslip.transportation_allowance = transport
            payslip.lunch_allowance = lunch
            payslip.employer_borne_deductions = employer_borne

            # calculate() recomputes gross_salary from components + runs deductions
            payslip.calculate()
            payslip.save()

            # Keep PayrollEntry in sync so next month's generation uses the correct gross
            entry = payslip.employee.payroll_entries.order_by('-updated_at').first()
            if entry:
                entry.basic = basic
                entry.housing = housing
                entry.transportation = transport
                entry.lunch = lunch
                entry.gross = Decimal(str(new_gross))
                entry.net = Decimal(str(payslip.net_salary))
                entry.employer_borne_deductions = employer_borne
                entry.save()

            updated += 1

        return Response({'updated': updated, 'employer_borne_deductions': employer_borne})

    @action(detail=False, methods=['get'])
    def remittance_report(self, request):
        """
        Get statutory remittance report for a period
        Aggregates PAYE (ZRA), NAPSA (employee + employer), and NHIMA (employee + employer)
        GET: ?year=2026&month=1
        """
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        
        if not year or not month:
            return Response(
                {'error': 'year and month query parameters required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        payslips = self.get_queryset().filter(
            period__year=year,
            period__month=month
        )
        
        # Aggregate statutory amounts
        total_paye = 0
        total_napsa_employee = 0
        total_napsa_employer = 0
        total_nhima_employee = 0
        total_nhima_employer = 0
        employee_breakdown = []
        
        for payslip in payslips:
            paye = float(payslip.paye_tax)
            napsa_emp = float(payslip.napsa_employee)
            napsa_empr = float(payslip.napsa_employer)
            nhima_emp = float(payslip.nhima_employee)
            nhima_empr = float(payslip.nhima_employer)
            
            total_paye += paye
            total_napsa_employee += napsa_emp
            total_napsa_employer += napsa_empr
            total_nhima_employee += nhima_emp
            total_nhima_employer += nhima_empr
            
            employee_breakdown.append({
                'employee_id': payslip.employee.employee_id,
                'employee_name': payslip.employee.full_name,
                'gross_salary': round(float(payslip.gross_salary), 2),
                'paye': round(paye, 2),
                'napsa_employee': round(napsa_emp, 2),
                'napsa_employer': round(napsa_empr, 2),
                'nhima_employee': round(nhima_emp, 2),
                'nhima_employer': round(nhima_empr, 2),
            })
        
        return Response({
            'period': {'year': year, 'month': month},
            'zra': {
                'paye_total': round(total_paye, 2),
            },
            'napsa': {
                'employee_total': round(total_napsa_employee, 2),
                'employer_total': round(total_napsa_employer, 2),
                'combined_total': round(total_napsa_employee + total_napsa_employer, 2),
            },
            'nhima': {
                'employee_total': round(total_nhima_employee, 2),
                'employer_total': round(total_nhima_employer, 2),
                'combined_total': round(total_nhima_employee + total_nhima_employer, 2),
            },
            'employee_breakdown': employee_breakdown,
            'summary': {
                'total_employees': len(payslips),
                'total_gross_payroll': round(sum(float(p.gross_salary) for p in payslips), 2),
                'total_statutory_deductions': round(
                    total_paye + total_napsa_employee + total_nhima_employee, 2
                ),
                'total_employer_contributions': round(
                    total_napsa_employer + total_nhima_employer, 2
                ),
            }
        })

    @action(detail=False, methods=['get'])
    def napsa_return(self, request):
        """Generate NAPSA monthly return in Excel format with proper formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        napsa_account = request.query_params.get('napsa_account', '11114862')  # Default or from settings

        if not year or not month:
            return Response({'error': 'year and month query parameters required'}, status=status.HTTP_400_BAD_REQUEST)

        payslips = self.get_queryset().filter(period__year=year, period__month=month)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"NAPSA Return {month}-{year}"
        
        # Header row with official NAPSA fields
        headers = [
            'NAPSA Account Number',
            'Year',
            'Month',
            'SSN',
            'NRC',
            'Surname',
            'First Name',
            'Other Names',
            'Date of Birth',
            'Gross Pay',
            'Employee Share',
            'Employer Share'
        ]
        
        # Write headers with formatting
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write employee data
        for p in payslips:
            emp = p.employee
            # Split name into surname and first name
            name_parts = emp.full_name.split()
            surname = name_parts[-1] if name_parts else ''
            first_name = name_parts[0] if len(name_parts) > 0 else ''
            other_names = ' '.join(name_parts[1:-1]) if len(name_parts) > 2 else ''
            
            # Format date of birth
            dob = emp.date_of_birth.strftime('%d/%m/%Y') if emp.date_of_birth else ''
            
            ws.append([
                napsa_account,
                year,
                month,
                emp.napsa_number or emp.sss_number,
                emp.nrc,
                surname.upper(),
                first_name.upper(),
                other_names.upper(),
                dob,
                float(p.gross_salary),
                float(p.napsa_employee),
                float(p.napsa_employer)
            ])
        
        # Set column widths for readability
        column_widths = [25, 8, 8, 20, 18, 20, 20, 20, 15, 15, 15, 15]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
        
        # Format currency columns
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=12):
            for cell in row:
                cell.number_format = '#,##0.00'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="napsa-return-{year}-{month}.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def nhima_return(self, request):
        """Generate NHIMA monthly return in Excel format with proper formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        nhima_account = request.query_params.get('nhima_account', '11114862')  # Default or from settings

        if not year or not month:
            return Response({'error': 'year and month query parameters required'}, status=status.HTTP_400_BAD_REQUEST)

        payslips = self.get_queryset().filter(period__year=year, period__month=month)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"NHIMA Return {month}-{year}"
        
        # Header row with official NHIMA fields
        headers = [
            'NHIMA Account Number',
            'Year',
            'Month',
            'NHIMA Number',
            'NRC',
            'Surname',
            'First Name',
            'Gross Pay',
            'Employee Contribution',
            'Employer Contribution'
        ]

        # Write headers with formatting
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write employee data
        for p in payslips:
            emp = p.employee
            # Split name into surname and first name
            name_parts = emp.full_name.split()
            surname = name_parts[-1] if name_parts else ''
            first_name = name_parts[0] if len(name_parts) > 0 else ''

            ws.append([
                nhima_account,
                year,
                month,
                emp.nhima or '',
                emp.nrc,
                surname.upper(),
                first_name.upper(),
                float(p.gross_salary),
                float(p.nhima_employee),
                float(p.nhima_employer)
            ])
        
        # Set column widths for readability
        column_widths = [25, 8, 8, 20, 18, 20, 20, 15, 20, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
        
        # Format currency columns
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
            for cell in row:
                cell.number_format = '#,##0.00'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="nhima-return-{year}-{month}.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export payslips for a period
        GET: ?year=2026&month=1
        """
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        
        if not year or not month:
            return Response(
                {'error': 'year and month query parameters required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        payslips = self.get_queryset().filter(
            period__year=year,
            period__month=month
        )
        
        serializer = PayslipSerializer(payslips, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def bank_transfer_export(self, request):
        """
        Export salary bank transfer file as Excel.
        GET: ?year=2026&month=1
        Columns: Employee ID, Employee Name, Bank, Branch, Account Number, Account Name, Net Salary, Currency
        """
        import openpyxl
        from io import BytesIO

        year = request.query_params.get('year')
        month = request.query_params.get('month')
        if not year or not month:
            return Response({'error': 'year and month are required'}, status=status.HTTP_400_BAD_REQUEST)

        payslips = self.get_queryset().filter(
            period__year=year,
            period__month=month,
        ).select_related('employee')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bank Transfer'

        headers = [
            'No.', 'Employee ID', 'Employee Name',
            'Bank Name', 'Branch', 'Account Number', 'Account Name',
            'Net Salary', 'Currency',
        ]
        ws.append(headers)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for idx, payslip in enumerate(payslips, start=1):
            emp = payslip.employee
            ws.append([
                idx,
                emp.employee_id,
                emp.full_name,
                emp.bank_name or '',
                emp.bank_branch or '',
                emp.bank_account_number or '',
                emp.bank_account_name or emp.full_name,
                float(payslip.net_salary or 0),
                'ZMW',
            ])

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        period_label = f"{year}_{str(month).zfill(2)}"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="BankTransfer_{period_label}.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def calculate_from_net(self, request):
        """
        Calculate gross salary from net salary using statutory settings.
        POST: { "net_salary": 5000, "period_id": 1 }
        """
        net_salary = request.data.get('net_salary')
        if not net_salary:
            return Response({'error': 'net_salary is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            net_salary = float(net_salary)
            
            # Get workspace from request if available
            workspace = None
            if hasattr(request.user, 'workspace_memberships'):
                membership = request.user.workspace_memberships.first()
                if membership:
                    workspace = membership.workspace
            
            result = calculate_gross_from_net(net_salary, workspace)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def preview_auto_data(self, request):
        """
        Preview auto-fetched leave, absenteeism, and double ticket data for a payslip
        POST: { "employee_id": 1, "year": 2026, "month": 2 }
        Returns: { "unpaid_leave_days": 3, "absenteeism_days": 1, "double_ticket_payment": 150.00, "absenteeism_deduction": 100.00 }
        """
        from apps.hcm.models import Employee
        from apps.leave.models import LeaveRequest, Absenteeism, DoubleTicketRequest
        from apps.payroll.models import PayrollEntry
        from apps.payroll.utils import calculate_unpaid_leave_deduction
        from datetime import date
        from calendar import monthrange
        
        employee_id = request.data.get('employee_id')
        year = request.data.get('year')
        month = request.data.get('month')
        
        if not all([employee_id, year, month]):
            return Response(
                {'error': 'employee_id, year, and month are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            employee = Employee.objects.get(id=employee_id)
            
            # Get or create a temporary period to use the auto-fetch methods
            period, _ = PayrollPeriod.objects.get_or_create(year=year, month=month)
            
            # Calculate month range
            first_day = date(year, month, 1)
            last_day = date(year, month, monthrange(year, month)[1])
            
            # Create a temporary payslip instance (not saved to DB)
            temp_payslip = Payslip(
                employee=employee,
                period=period,
                basic_salary=0  # Just for initialization
            )
            
            # Use the auto-fetch methods
            unpaid_days = temp_payslip.auto_fetch_unpaid_leave_days()
            absenteeism_days = temp_payslip.auto_fetch_absenteeism_days()
            double_ticket_payment = temp_payslip.auto_fetch_double_ticket_payment()
            
            # Calculate deductions based on basic salary
            absenteeism_deduction = 0
            unpaid_leave_deduction = 0
            payroll_entry = PayrollEntry.objects.filter(employee=employee).first()
            
            if payroll_entry:
                basic_salary = float(payroll_entry.basic)
                gross_salary = float(payroll_entry.gross) if payroll_entry.gross else basic_salary
                
                # Absenteeism deduction: (basic_salary / 26) * absenteeism_days
                if basic_salary > 0 and absenteeism_days > 0:
                    daily_rate = basic_salary / 26
                    absenteeism_deduction = round(daily_rate * float(absenteeism_days), 2)
                
                # Unpaid leave deduction: (gross_salary / 26) * unpaid_days
                if gross_salary > 0 and unpaid_days > 0:
                    unpaid_leave_deduction = calculate_unpaid_leave_deduction(gross_salary, unpaid_days)
            
            # DEBUG: Query all leave data for this employee in this period
            all_leaves = LeaveRequest.objects.filter(
                employee=employee,
                start_date__lte=last_day,
                end_date__gte=first_day
            ).values('id', 'leave_type', 'status', 'start_date', 'end_date', 'days')
            
            all_absenteeism = Absenteeism.objects.filter(
                employee=employee,
                date__gte=first_day,
                date__lte=last_day
            ).values('id', 'date', 'status', 'reason_provided')
            
            all_double_tickets = DoubleTicketRequest.objects.filter(
                employee=employee,
                work_date__gte=first_day,
                work_date__lte=last_day
            ).values('id', 'work_date', 'status', 'hours_worked', 'calculated_amount')
            
            return Response({
                'employee_name': employee.full_name,
                'period': f"{period.get_month_display()} {period.year}",
                'unpaid_leave_days': unpaid_days,
                'unpaid_leave_deduction': unpaid_leave_deduction,
                'absenteeism_days': absenteeism_days,
                'absenteeism_deduction': absenteeism_deduction,
                'double_ticket_payment': round(double_ticket_payment, 2),
                'message': 'These values will be automatically applied when creating the payslip',
                'debug_info': {
                    'note': 'Only UNPAID approved leaves, UNJUSTIFIED absenteeism, and APPROVED/PAID double tickets are auto-fetched',
                    'all_leaves_in_period': list(all_leaves),
                    'all_absenteeism_in_period': list(all_absenteeism),
                    'all_double_tickets_in_period': list(all_double_tickets),
                }
            })
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def adjust_basic_salary(self, request, pk=None):
        """
        Adjust basic salary and recalculate payslip.
        POST: { "basic_salary": 15000 }
        """
        payslip = self.get_object()
        new_basic = request.data.get('basic_salary')
        
        if not new_basic:
            return Response({'error': 'basic_salary is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            old_basic = str(payslip.basic_salary)
            payslip.basic_salary = float(new_basic)
            payslip.calculate()
            payslip.save()
            
            # Log the change
            PayslipAuditLog.objects.create(
                payslip=payslip,
                action='UPDATED',
                user=request.user,
                changes={
                    'field': 'basic_salary',
                    'old_value': old_basic,
                    'new_value': str(payslip.basic_salary),
                    'old_net': old_basic,
                    'new_net': str(payslip.net_salary),
                }
            )
            
            serializer = PayslipSerializer(payslip)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def apply_auto_data(self, request, pk=None):
        """Refresh and persist unpaid leave, absenteeism, and double ticket values for this payslip."""
        payslip = self.get_object()

        try:
            old_values = {
                'unpaid_leave_days': str(payslip.unpaid_leave_days),
                'unpaid_leave_deduction': str(payslip.unpaid_leave_deduction),
                'absenteeism_days': str(payslip.absenteeism_days),
                'absenteeism_deduction': str(payslip.absenteeism_deduction),
                'double_ticket_payment': str(payslip.double_ticket_payment),
                'gross_salary': str(payslip.gross_salary),
                'net_salary': str(payslip.net_salary),
            }

            auto_unpaid_days = payslip.auto_fetch_unpaid_leave_days()
            auto_absenteeism_days = payslip.auto_fetch_absenteeism_days()
            auto_double_ticket = payslip.auto_fetch_double_ticket_payment()

            payslip.unpaid_leave_days = auto_unpaid_days
            payslip.absenteeism_days = auto_absenteeism_days
            payslip.double_ticket_payment = auto_double_ticket

            # Recompute these deductions from refreshed day counts
            payslip.unpaid_leave_deduction = 0
            payslip.absenteeism_deduction = 0

            payslip.calculate()
            payslip.save()

            new_values = {
                'unpaid_leave_days': str(payslip.unpaid_leave_days),
                'unpaid_leave_deduction': str(payslip.unpaid_leave_deduction),
                'absenteeism_days': str(payslip.absenteeism_days),
                'absenteeism_deduction': str(payslip.absenteeism_deduction),
                'double_ticket_payment': str(payslip.double_ticket_payment),
                'gross_salary': str(payslip.gross_salary),
                'net_salary': str(payslip.net_salary),
            }

            PayslipAuditLog.objects.create(
                payslip=payslip,
                action='UPDATED',
                user=request.user,
                changes={'auto_refresh': {'old': old_values, 'new': new_values}}
            )

            serializer = PayslipSerializer(payslip)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """
        Download payslip as PDF receipt
        """
        try:
            payslip = self.get_object()
            
            # Generate PDF with workspace name if available
            workspace_name = None
            if hasattr(request, 'workspace') and request.workspace:
                workspace_name = request.workspace.name
            pdf = generate_payslip_pdf(payslip, workspace_name=workspace_name)
            
            # Create HTTP response with PDF
            response = HttpResponse(pdf, content_type='application/pdf')
            period_label = str(payslip.period).replace(" ", "_")
            filename = f'payslip_{payslip.employee.employee_id}_{period_label}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class PAYEReturnViewSet(viewsets.ModelViewSet):
    """
    ViewSet for PAYE (Pay As You Earn) statutory returns.
    Provides endpoints to generate, view, and manage PAYE returns by period.
    """
    serializer_class = None  # Will be set in __init__
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['period', 'status']
    ordering_fields = ['period__year', 'period__month', 'created_at']
    ordering = ['-period__year', '-period__month']
    
    def get_queryset(self):
        """Filter PAYE returns by workspace."""
        if hasattr(self.request, 'workspace') and self.request.workspace:
            return PAYEReturn.objects.filter(
                workspace=self.request.workspace
            ).select_related('period', 'workspace', 'created_by')
        return PAYEReturn.objects.none()
    
    def get_serializer_class(self):
        """Use appropriate serializer based on action."""
        from .serializers import PAYEReturnSerializer
        if self.action == 'retrieve_detail':
            from .serializers import PAYEReturnDetailRowSerializer
            return PAYEReturnDetailRowSerializer
        return PAYEReturnSerializer
    
    def perform_create(self, serializer):
        """Automatically set workspace and user."""
        workspace = getattr(self.request, 'workspace', None)
        serializer.save(workspace=workspace, created_by=self.request.user)
    
    @action(detail=False, methods=['post'])
    def generate(self, request):
        """
        Generate a PAYE return for a given period.
        POST: { "period": 1 }  (period ID)
        """
        from .models import PAYEReturn
        from .serializers import PAYEReturnSerializer
        
        period_id = request.data.get('period')
        if not period_id:
            return Response(
                {'error': 'period ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            period = PayrollPeriod.objects.get(id=period_id)
            workspace = getattr(request, 'workspace', None)
            
            if not workspace:
                return Response(
                    {'error': 'Workspace context is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get or create PAYE return for this period and workspace
            paye_return, created = PAYEReturn.objects.get_or_create(
                workspace=workspace,
                period=period,
                defaults={'created_by': request.user}
            )
            
            # Calculate the return
            paye_return.calculate()
            
            serializer = PAYEReturnSerializer(paye_return)
            status_code = status.HTTP_201_CREATED if created else status.HTTP_200_OK
            return Response(serializer.data, status=status_code)
        
        except PayrollPeriod.DoesNotExist:
            return Response(
                {'error': 'Period not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def detail_rows(self, request, pk=None):
        """
        Get detailed PAYE return rows (one per employee) with all columns.
        Returns CSV-friendly list of rows.
        """
        from .serializers import PAYEReturnDetailRowSerializer
        
        try:
            paye_return = self.get_object()
            rows = paye_return.generate_detailed_rows()
            
            serializer = PAYEReturnDetailRowSerializer(rows, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def export_csv(self, request, pk=None):
        """
        Export PAYE return as CSV file.
        Columns: TPIN, Full Name, Employee Nature, Gross Pay, Chargeable Amount, Tax Credit, Tax Payable
        """
        try:
            paye_return = self.get_object()
            rows = paye_return.generate_detailed_rows()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            period_label = f"{paye_return.period.get_month_display()}_{paye_return.period.year}"
            response['Content-Disposition'] = f'attachment; filename="paye-return-{period_label}.csv"'
            
            writer = csv.DictWriter(response, fieldnames=[
                'tpin', 'full_name', 'employee_nature', 'gross_pay',
                'chargeable_amount', 'tax_credit', 'tax_payable', 'employee_id'
            ])
            
            writer.writeheader()
            for row in rows:
                writer.writerow({
                    'tpin': row['tpin'],
                    'full_name': row['full_name'],
                    'employee_nature': row['employee_nature'],
                    'gross_pay': row['gross_pay'],
                    'chargeable_amount': row['chargeable_amount'],
                    'tax_credit': row['tax_credit'],
                    'tax_payable': row['tax_payable'],
                    'employee_id': row['employee_id'],
                })
            
            return response
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def mark_submitted(self, request, pk=None):
        """
        Mark PAYE return as submitted to tax authority.
        POST: optional { "submission_notes": "..." }
        """
        from .serializers import PAYEReturnSerializer
        
        try:
            paye_return = self.get_object()
            paye_return.status = 'SUBMITTED'
            paye_return.submitted_at = timezone.now()
            paye_return.submission_notes = request.data.get('submission_notes', '')
            paye_return.save()

            serializer = PAYEReturnSerializer(paye_return)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class WorkspaceStatutorySettingsViewSet(viewsets.ModelViewSet):
    serializer_class = WorkspaceStatutorySettingsSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['workspace']

    def get_queryset(self):
        qs = WorkspaceStatutorySettings.objects.select_related('workspace')
        workspace_id = self.request.query_params.get('workspace')
        if workspace_id:
            qs = qs.filter(workspace_id=workspace_id)
        elif hasattr(self.request, 'workspace') and self.request.workspace:
            qs = qs.filter(workspace=self.request.workspace)
        return qs


class PayeTaxBandViewSet(viewsets.ModelViewSet):
    serializer_class = PayeTaxBandSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['workspace']
    ordering = ['order', 'min_amount']

    def get_queryset(self):
        qs = PayeTaxBand.objects.select_related('workspace')
        workspace_id = self.request.query_params.get('workspace')
        if workspace_id:
            qs = qs.filter(workspace_id=workspace_id)
        elif hasattr(self.request, 'workspace') and self.request.workspace:
            qs = qs.filter(workspace=self.request.workspace)
        return qs


class ComplianceDocumentViewSet(viewsets.ModelViewSet):
    serializer_class = ComplianceDocumentSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['workspace', 'document_type']
    ordering = ['document_type']

    def get_queryset(self):
        qs = ComplianceDocument.objects.select_related('workspace')
        workspace_id = self.request.query_params.get('workspace')
        if workspace_id:
            qs = qs.filter(workspace_id=workspace_id)
        elif hasattr(self.request, 'workspace') and self.request.workspace:
            qs = qs.filter(workspace=self.request.workspace)
        return qs

    def perform_create(self, serializer):
        workspace = getattr(self.request, 'workspace', None)
        if workspace and 'workspace' not in serializer.validated_data:
            serializer.save(workspace=workspace)
        else:
            serializer.save()

    @action(detail=False, methods=['get'])
    def summary(self, request):
        workspace_id = request.query_params.get('workspace')
        if workspace_id:
            qs = ComplianceDocument.objects.filter(workspace_id=workspace_id)
        elif hasattr(request, 'workspace') and request.workspace:
            qs = ComplianceDocument.objects.filter(workspace=request.workspace)
        else:
            return Response({'error': 'Workspace context required'}, status=status.HTTP_400_BAD_REQUEST)

        docs = list(qs)
        total = len(docs)
        active = sum(1 for d in docs if d.computed_status == 'Active')
        permanent = sum(1 for d in docs if d.computed_status == 'Permanent')
        expiring_soon = sum(1 for d in docs if d.computed_status == 'Expiring Soon')
        expired = sum(1 for d in docs if d.computed_status == 'Expired')

        valid = active + permanent
        compliance_pct = round((valid / total) * 100) if total > 0 else 0

        serializer = ComplianceDocumentSerializer(docs, many=True)
        return Response({
            'total': total,
            'active': active,
            'permanent': permanent,
            'expiring_soon': expiring_soon,
            'expired': expired,
            'compliance_pct': compliance_pct,
            'documents': serializer.data,
        })
